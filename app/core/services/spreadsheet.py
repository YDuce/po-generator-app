"""Spreadsheet builder utilities.

Layer: core
"""

from pathlib import Path
import csv
from typing import List, Dict
from openpyxl import load_workbook

__all__ = ["SpreadsheetBuilder"]


class SpreadsheetBuilder:
    """Build spreadsheets from row dictionaries using templates."""

    def __init__(self, template_path: Path):
        self.template_path = template_path

    def write_excel(self, rows: List[Dict[str, object]], output_path: Path) -> Path:
        wb = load_workbook(self.template_path)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header))
        wb.save(output_path)
        return output_path

    def write_csv(self, rows: List[Dict[str, object]], output_path: Path) -> Path:
        if not rows:
            raise ValueError("No data to write")
        headers = rows[0].keys()
        with open(output_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
        return output_path
